from openpyxl import Workbook
from openpyxl.styles import Font


def build_summary_excel(dest_path, summary):
    wb = Workbook()

    overview = wb.active
    overview.title = "Overview"
    overview.append(["Metric", "Value"])
    overview["A1"].font = overview["B1"].font = Font(bold=True)
    overview.append(["Active Policies", summary["active_policies"]])
    overview.append(["Expired Policies", summary["expired_policies"]])
    overview.append(["Cancelled Policies", summary["cancelled_policies"]])
    overview.append(["Total Customers", summary["total_customers"]])
    overview.append(["Premium Collected", summary["premium_collected"]])
    overview.append(["Pending Claims", summary["claim_stats"]["pending"]])
    overview.append(["Approved Claims", summary["claim_stats"]["approved"]])
    overview.append(["Rejected Claims", summary["claim_stats"]["rejected"]])
    filters = summary.get("filters", {})
    overview.append(["Period", f"{filters.get('start_date', '-')} to {filters.get('end_date', '-')}"])
    overview.append(["Policy type filter", filters.get("policy_type") or "All"])

    policy_type_sheet = wb.create_sheet("Policies by Type")
    policy_type_sheet.append(["Policy Type", "Count"])
    for row in summary.get("policy_type_breakdown", []):
        policy_type_sheet.append([row["policy_type"], row["count"]])

    premiums_sheet = wb.create_sheet("Monthly Premiums")
    premiums_sheet.append(["Year", "Month", "Total Collected"])
    for row in summary["monthly_premium_collection"]:
        premiums_sheet.append([row["year"], row["month"], row["total"]])

    growth_sheet = wb.create_sheet("Customer Growth")
    growth_sheet.append(["Year", "Month", "New Customers"])
    for row in summary["customer_growth"]:
        growth_sheet.append([row["year"], row["month"], row["count"]])

    wb.save(dest_path)
    return dest_path
